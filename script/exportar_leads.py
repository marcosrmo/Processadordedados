#!/usr/bin/env python3
"""
exportar_leads.py — Processamento local de planilhas DataForge

Como usar:
  1. Coloque seus arquivos CSV, XLSX, XLS ou TXT na pasta  input/
  2. Execute:  python script/exportar_leads.py
  3. A planilha gerada fica em:  exports/leads_YYYY-MM-DD_HH-MM-SS.xlsx

Filtros aplicados automaticamente:
  - Sem Engano              : exclui registros com a palavra "engano"
  - Sem Falecido            : exclui registros com a palavra "falecido"
  - Somente com Telefone    : exclui registros sem telefone preenchido
  - Sem duplicatas          : mantém apenas um registro por número de telefone
  - Dígito 9 automático     : adiciona o 9 em celulares com 8 dígitos locais

Não precisa de banco de dados — processa os arquivos diretamente.
"""

import os
import sys
import re
import csv
import unicodedata
from datetime import datetime
from pathlib import Path

# ── Dependências ───────────────────────────────────────────────────────────────
try:
    import chardet
except ImportError:
    print("ERRO: chardet não instalado. Execute: pip install chardet")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERRO: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("ERRO: pandas não instalado. Execute: pip install pandas")
    sys.exit(1)


# ── Pastas ────────────────────────────────────────────────────────────────────
ROOT_DIR   = Path(__file__).resolve().parent.parent
INPUT_DIR  = ROOT_DIR / "input"
OUTPUT_DIR = ROOT_DIR / "exports"

INPUT_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
#  MAPA DE CAMPOS — mesmos aliases do FIELD_MAP em server/storage.ts
# ══════════════════════════════════════════════════════════════════════════════

FIELD_MAP = {
    "name": [
        "nome", "nomecompleto", "nome completo", "name", "cliente", "nome cliente",
        "nome do cliente", "comprador", "proprietario", "segurado", "paciente",
        "funcionario", "colaborador", "consumidor", "pessoa", "razao social",
        "titular", "socio", "representante", "responsavel", "diretor", "vendedor",
    ],
    "cpf": [
        "cpf", "cpf do cliente", "cpf_cliente", "cpfcliente", "cpf cliente",
    ],
    "cnpj": [
        "cnpj", "cnpjempresa", "cnpj empresa", "cgc", "cnpj do cliente",
    ],
    "cpfCnpj": [
        "cpf cnpj", "cpfcnpj", "documento", "doc", "cpf ou cnpj",
        "cpf/cnpj", "cpfoucnpj", "cpf_cnpj",
    ],
    "phone": [
        "telefone", "tel", "fone", "phone", "ramal", "fax",
        "contato", "tel 1", "telefone 1", "numero",
        "telefone fixo", "tel fixo", "fixo", "residencial",
        "telefone do cliente", "tel do cliente", "fone do cliente",
        "telefone cliente", "tel cliente", "nr telefone", "num telefone",
        "numero telefone", "telefone comercial", "tel comercial",
    ],
    "phone2": [
        "celular", "cel", "movel", "mobile", "telefonecelular", "telcel",
        "fonecelular", "telefonemovel", "fonemovel", "celulartel",
        "whatsapp", "wpp", "zap", "telefone celular", "tel celular",
        "numero celular", "celular1", "cel1", "telefone2", "tel2",
        "fone2", "phone2", "tel 2", "telefone 2", "celular2",
        "cel2", "whatsapp2", "fone 2", "contato2",
    ],
    "phone3": [
        "telefone3", "tel3", "fone3", "phone3", "tel 3", "telefone 3",
        "celular3", "cel3", "whatsapp3", "fone 3", "contato3",
    ],
    "phone4": [
        "telefone4", "tel4", "fone4", "phone4", "tel 4", "telefone 4",
        "celular4", "cel4", "whatsapp4", "fone 4", "contato4",
    ],
    "ddd": [
        "ddd", "dd", "cod area", "codigo area", "codigo de area",
        "area code", "areacode", "prefixo", "cod. area", "cod ddd",
    ],
    "email": [
        "email", "e-mail", "mail", "correio", "emailaddress", "email address",
        "e mail", "emailcliente", "email cliente", "email do cliente",
    ],
    "address": [
        "endereco", "logradouro", "address", "rua", "r.", "avenida", "av", "av.",
        "alameda", "travessa", "estrada", "rodovia", "end", "logr",
    ],
    "number": ["numero", "nro", "num", "number", "no"],
    "complement": [
        "complemento", "compl", "comp", "apto", "apartamento", "casa", "bloco", "sala", "apt",
    ],
    "neighborhood": ["bairro", "district", "neighborhood"],
    "city": ["cidade", "municipio", "localidade", "city"],
    "state": ["estado", "uf", "sigla", "est", "provincia", "state"],
    "cep": ["cep", "zip", "zipcode", "postalcode", "codigo postal"],
    "ticketAverage": [
        "ticket medio", "ticketmedio", "ticket", "faturamento",
        "valor", "valor compra", "valor total", "total", "preco", "price",
        "valor pedido", "valor do pedido", "valor nota", "valor de nota",
        "valor nf", "vl pedido", "vl total", "vl nota", "vlpedido", "vltotal",
        "valornota", "valorpedido", "valorcompra", "valortotal",
        "total pedido", "total nota", "nf valor", "nota fiscal valor",
        "preco total", "preco unitario", "valor unitario", "vl unitario",
        "maior compra", "melhor compra", "vl compra", "valor da compra",
    ],
    "purchaseDate": [
        "data compra", "datacompra", "data", "dt", "data venda", "datavenda",
        "ultima compra", "data_venda", "date", "dataemissao", "data emissao",
        "dt. ult. compra", "dt ult compra", "dtultcompra",
        "data pedido", "datapedido", "dt pedido", "dtpedido",
        "data saida", "datasaida", "dt saida", "dtsaida", "data de saida",
        "data faturamento", "datafaturamento", "dt faturamento", "dtfaturamento",
        "data de faturamento", "data nota", "data nf", "datanf",
        "data emissao nota", "data lancamento", "data entrada",
        "data movimento", "data operacao", "data transacao",
    ],
    "produto": [
        "produto", "product", "mercadoria", "item", "descricao produto",
        "descricaoproduto", "nome produto", "nomeproduto", "servico",
        "descricao", "desc produto", "desc item",
    ],
    "quantidade": [
        "quantidade", "qtd", "qt", "qtde", "qty", "quant", "qnt",
        "qtdade", "qtidade",
    ],
}

# Índice invertido: alias normalizado → nome do campo
_ALIAS_INDEX: dict[str, str] = {}
for _field, _aliases in FIELD_MAP.items():
    for _alias in _aliases:
        _ALIAS_INDEX[_alias] = _field


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def normalize_str(s: str) -> str:
    """Remove acentos, coloca em minúsculas, remove caracteres especiais."""
    s = s.lower().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9 ]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def only_digits(v) -> str:
    """Remove tudo que não é dígito."""
    if v is None or (isinstance(v, float) and str(v) == "nan"):
        return ""
    return re.sub(r"\D", "", str(v))


def prepend_ddd(phone: str, ddd: str) -> str:
    """Concatena DDD apenas se o número tiver menos de 10 dígitos."""
    if not phone or not ddd:
        return phone
    if len(phone) >= 10:
        return phone  # DDD já presente — não duplica
    return ddd + phone


def add_nine_digit(phone: str) -> str:
    """
    Adiciona o 9º dígito em celulares brasileiros com 10 dígitos (DDD + 8).
    Só adiciona se o primeiro dígito do número local for 6, 7, 8 ou 9.
    """
    if len(phone) != 10:
        return phone
    ddd_part   = phone[:2]
    local_part = phone[2:]
    if local_part and local_part[0] in "6789":
        return ddd_part + "9" + local_part
    return phone


def map_columns(headers: list[str]) -> dict[str, str]:
    """
    Recebe lista de cabeçalhos originais e retorna mapeamento
    { header_original → nome_do_campo }.

    Estratégias em ordem de prioridade:
    1. Match exato: normalized_header == alias
    2. Match por prefixo: normalized_header começa com alias (ex: "telefone do cliente" → "telefone")
    3. Match por contenção: alias está contido em normalized_header
    """
    mapping: dict[str, str] = {}

    # Pré-ordena aliases do mais longo ao mais curto para priorizar matches mais específicos
    sorted_aliases = sorted(_ALIAS_INDEX.items(), key=lambda x: len(x[0]), reverse=True)

    for h in headers:
        norm = normalize_str(h)

        # 1. Match exato
        if norm in _ALIAS_INDEX:
            field = _ALIAS_INDEX[norm]
            if field not in mapping.values():
                mapping[h] = field
            continue

        # 2. Match por prefixo (ex: "telefone do cliente" começa com "telefone")
        best_field  = None
        best_len    = 0
        for alias, field in sorted_aliases:
            if norm.startswith(alias) and len(alias) > best_len:
                best_field = field
                best_len   = len(alias)

        if best_field and best_field not in mapping.values():
            mapping[h] = best_field
            continue

        # 3. Match por contenção (ex: "nr telefone fixo" contém "telefone fixo")
        for alias, field in sorted_aliases:
            if len(alias) >= 3 and alias in norm:
                if field not in mapping.values():
                    mapping[h] = field
                break

    return mapping


def split_cpf_cnpj(raw) -> tuple[str, str]:
    """Decide se valor misto é CPF (11 dígitos) ou CNPJ (14 dígitos)."""
    if not raw:
        return ("", "")
    digits = only_digits(raw)
    if len(digits) == 11:
        return (digits, "")
    if len(digits) == 14:
        return ("", digits)
    if 0 < len(digits) < 12:
        return (digits, "")
    return ("", digits)


def detect_encoding(path: Path) -> str:
    raw = path.read_bytes()[:50_000]
    result = chardet.detect(raw)
    enc = result.get("encoding") or "utf-8"
    # latin1 é seguro como fallback universal para português
    if enc.lower() in ("ascii", "windows-1252", "iso-8859-1"):
        enc = "latin1"
    return enc


def detect_delimiter(first_line: str) -> str:
    counts = {
        ",":  first_line.count(","),
        ";":  first_line.count(";"),
        "|":  first_line.count("|"),
        "\t": first_line.count("\t"),
    }
    return max(counts, key=counts.get)


# ══════════════════════════════════════════════════════════════════════════════
#  LEITURA DE ARQUIVOS
# ══════════════════════════════════════════════════════════════════════════════

def read_file(path: Path) -> list[dict]:
    """
    Lê CSV, TXT, XLSX ou XLS e retorna lista de dicts com os dados brutos.
    """
    ext = path.suffix.lower()

    if ext in (".xlsx",):
        try:
            df = pd.read_excel(path, engine="openpyxl", dtype=str, header=0)
            df = df.fillna("")
            return df.to_dict(orient="records")
        except Exception as e:
            print(f"  ERRO ao ler {path.name}: {e}")
            return []

    if ext in (".xls",):
        try:
            df = pd.read_excel(path, engine="xlrd", dtype=str, header=0)
            df = df.fillna("")
            return df.to_dict(orient="records")
        except Exception as e:
            print(f"  ERRO ao ler {path.name}: {e}")
            return []

    if ext in (".csv", ".txt"):
        enc = detect_encoding(path)
        try:
            text = path.read_bytes().decode(enc, errors="replace")
        except Exception:
            text = path.read_bytes().decode("latin1", errors="replace")

        # Remove BOM se existir
        if text.startswith("\ufeff"):
            text = text[1:]

        lines = text.splitlines()
        if not lines:
            return []

        delimiter = detect_delimiter(lines[0])
        reader = csv.DictReader(
            lines,
            delimiter=delimiter,
            quotechar='"',
            skipinitialspace=True,
        )
        rows = []
        for row in reader:
            rows.append({k: (v.strip() if v else "") for k, v in row.items()})
        return rows

    print(f"  Formato não suportado: {path.suffix} — ignorando {path.name}")
    return []


# ══════════════════════════════════════════════════════════════════════════════
#  PROCESSAMENTO DE UM ARQUIVO
# ══════════════════════════════════════════════════════════════════════════════

def process_rows(raw_rows: list[dict]) -> list[dict]:
    """
    Recebe as linhas brutas de um arquivo e retorna registros normalizados
    com os campos do FIELD_MAP.
    """
    if not raw_rows:
        return []

    # Mapear cabeçalhos
    headers    = list(raw_rows[0].keys())
    col_map    = map_columns(headers)  # { header_original → field_name }

    records = []
    for row in raw_rows:
        # Montar um dict field_name → valor para cada linha
        fields: dict[str, str] = {}
        for header, value in row.items():
            field = col_map.get(header)
            if field and field not in fields:
                v = str(value).strip() if value else ""
                if v:
                    fields[field] = v

        # CPF / CNPJ (campo misto)
        if "cpfCnpj" in fields and not ("cpf" in fields and "cnpj" in fields):
            cpf_v, cnpj_v = split_cpf_cnpj(fields.pop("cpfCnpj"))
            if cpf_v and "cpf" not in fields:
                fields["cpf"] = cpf_v
            if cnpj_v and "cnpj" not in fields:
                fields["cnpj"] = cnpj_v
        elif "cpfCnpj" in fields:
            fields.pop("cpfCnpj")

        # Normalizar CPF / CNPJ (só dígitos)
        if "cpf"  in fields: fields["cpf"]  = only_digits(fields["cpf"])
        if "cnpj" in fields: fields["cnpj"] = only_digits(fields["cnpj"])

        # Extrair DDD separado
        raw_ddd = ""
        if "ddd" in fields:
            ddd_digits = only_digits(fields["ddd"])
            raw_ddd    = ddd_digits[-2:] if len(ddd_digits) >= 2 else ""
            fields["ddd"] = raw_ddd

        # Normalizar telefones (só dígitos) + concatenar DDD
        for ph in ("phone", "phone2", "phone3", "phone4"):
            if ph in fields:
                digits = only_digits(fields[ph])
                digits = prepend_ddd(digits, raw_ddd)
                fields[ph] = digits if digits else ""
            else:
                fields[ph] = ""

        # Garante campos padrão vazios
        for f in ("name", "cpf", "cnpj", "email", "address", "number",
                  "complement", "neighborhood", "city", "state", "cep",
                  "ticketAverage", "purchaseDate", "produto", "quantidade", "ddd"):
            fields.setdefault(f, "")

        records.append(fields)

    return records


# ══════════════════════════════════════════════════════════════════════════════
#  MERGE / DEDUPLICAÇÃO NO PROCESSAMENTO
# ══════════════════════════════════════════════════════════════════════════════

def filled_count(rec: dict) -> int:
    return sum(1 for v in rec.values() if v)


def normalize_name(name: str) -> str:
    return normalize_str(name)


def name_similarity(a: str, b: str) -> float:
    na = normalize_name(a)
    nb = normalize_name(b)
    if not na or not nb:
        return 0.0
    tok_a = [t for t in na.split() if len(t) > 2]
    tok_b = [t for t in nb.split() if len(t) > 2]
    if not tok_a or not tok_b:
        return 0.0
    common = len([t for t in tok_a if t in tok_b])
    return common / max(len(tok_a), len(tok_b))


def canonical_phone(phone: str) -> str:
    """Remove o 9º dígito de celulares com 11 dígitos para comparação."""
    if len(phone) == 11 and phone[2] == "9":
        return phone[:2] + phone[3:]
    return phone


def merge_records(all_records: list[dict]) -> list[dict]:
    """
    Mescla registros com mesmo CPF, CNPJ ou telefone+nome similar.
    Mantém o registro mais completo como base, preenchendo campos vazios.
    """
    merged_list: list[dict] = []
    cpf_idx:   dict[str, int] = {}
    cnpj_idx:  dict[str, int] = {}
    phone_idx: dict[str, int] = {}

    def merge_into(dst: dict, src: dict) -> dict:
        result = dict(dst)
        for k, v in src.items():
            if not result.get(k) and v:
                result[k] = v
        return result

    for rec in all_records:
        match_idx = None

        cpf = rec.get("cpf", "")
        if cpf and cpf in cpf_idx:
            match_idx = cpf_idx[cpf]

        if match_idx is None:
            cnpj = rec.get("cnpj", "")
            if cnpj and cnpj in cnpj_idx:
                match_idx = cnpj_idx[cnpj]

        if match_idx is None:
            phone = rec.get("phone", "")
            if phone:
                cp = canonical_phone(phone)
                if cp in phone_idx:
                    idx = phone_idx[cp]
                    existing = merged_list[idx]
                    sim = name_similarity(rec.get("name", ""), existing.get("name", ""))
                    has_name     = bool(rec.get("name", "").strip())
                    existing_has = bool(existing.get("name", "").strip())
                    if (not has_name and not existing_has) or sim >= 0.7:
                        match_idx = idx

        if match_idx is not None:
            merged_list[match_idx] = merge_into(merged_list[match_idx], rec)
        else:
            new_idx = len(merged_list)
            merged_list.append(rec)

            cpf   = rec.get("cpf", "")
            cnpj  = rec.get("cnpj", "")
            phone = rec.get("phone", "")
            if cpf:   cpf_idx[cpf]   = new_idx
            if cnpj:  cnpj_idx[cnpj] = new_idx
            if phone:
                cp = canonical_phone(phone)
                if cp: phone_idx[cp] = new_idx

    return merged_list


# ══════════════════════════════════════════════════════════════════════════════
#  PERGUNTAS INTERATIVAS DE FILTRO
# ══════════════════════════════════════════════════════════════════════════════

def perguntar(pergunta: str, padrao: bool = True) -> bool:
    """
    Exibe uma pergunta s/n no terminal e retorna True ou False.
    O valor padrão é aplicado se o usuário pressionar Enter sem digitar nada.
    """
    opcoes = "[S/n]" if padrao else "[s/N]"
    while True:
        try:
            resp = input(f"  {pergunta} {opcoes}: ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            print()
            return padrao
        if resp == "":
            return padrao
        if resp in ("s", "sim", "y", "yes"):
            return True
        if resp in ("n", "nao", "não", "no"):
            return False
        print("    Digite  s  para Sim  ou  n  para Não.")


def perguntar_filtros() -> dict:
    """
    Pergunta ao usuário quais filtros deseja aplicar e retorna um dict
    com as opções escolhidas.
    """
    SEP = "-" * 60
    print(f"\n{SEP}")
    print("  CONFIGURAÇÃO DOS FILTROS")
    print(f"{SEP}")
    print("  Responda s (sim) ou n (não) para cada filtro.")
    print("  Pressionar Enter aceita a opção em MAIÚSCULA.\n")

    opts: dict[str, bool] = {}

    opts["sem_engano"] = perguntar(
        "Remover registros com a palavra 'engano'?", padrao=True
    )
    opts["sem_falecido"] = perguntar(
        "Remover registros com a palavra 'falecido'?", padrao=True
    )
    opts["somente_com_tel"] = perguntar(
        "Manter somente registros COM telefone preenchido?", padrao=True
    )

    print()
    print("  -- Deduplicação de telefone (pode ativar os dois juntos) --")
    opts["dup_exato"] = perguntar(
        "Remover duplicatas por telefone EXATO (dígitos idênticos)?", padrao=True
    )
    opts["dup_nove"] = perguntar(
        "Remover duplicatas considerando 9º dígito (11 9xxxx = 11 xxxx)?", padrao=True
    )

    print()
    opts["nome_obrigatorio"] = perguntar(
        "Manter somente registros COM nome preenchido?", padrao=False
    )
    opts["adicionar_nove"] = perguntar(
        "Adicionar o 9º dígito em celulares com 8 dígitos locais?", padrao=True
    )

    print(f"\n{SEP}")
    print("  RESUMO DOS FILTROS ESCOLHIDOS")
    print(f"{SEP}")
    labels = {
        "sem_engano":       "Remover 'engano'                    ",
        "sem_falecido":     "Remover 'falecido'                  ",
        "somente_com_tel":  "Somente com telefone                ",
        "dup_exato":        "Sem duplicatas (telefone exato)     ",
        "dup_nove":         "Sem duplicatas (com 9º dígito)      ",
        "nome_obrigatorio": "Nome obrigatório                    ",
        "adicionar_nove":   "Adicionar 9º dígito                 ",
    }
    for key, label in labels.items():
        status = "✔  SIM" if opts[key] else "✘  NÃO"
        print(f"  {label}: {status}")
    print(f"{SEP}\n")

    return opts


# ══════════════════════════════════════════════════════════════════════════════
#  FILTROS DE EXPORTAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

ENGANO_RE   = re.compile(r"\bengano\b",   re.IGNORECASE)
FALECIDO_RE = re.compile(r"\bfalecido\b", re.IGNORECASE)


def row_full_text(rec: dict) -> str:
    return " ".join(str(v) for v in rec.values() if v)


def apply_filters(records: list[dict], opts: dict) -> list[dict]:
    result = []

    # Conjuntos separados para cada modo de deduplicação
    seen_exact:     set[str] = set()   # dígitos crus, sem normalização
    seen_canonical: set[str] = set()   # canonical: remove o 9 de 11 dígitos

    count_engano   = 0
    count_falecido = 0
    count_sem_tel  = 0
    count_sem_nome = 0
    count_dup_exato = 0
    count_dup_nove  = 0

    for rec in records:
        text = row_full_text(rec)

        # Filtro: sem engano
        if opts.get("sem_engano") and ENGANO_RE.search(text):
            count_engano += 1
            continue

        # Filtro: sem falecido
        if opts.get("sem_falecido") and FALECIDO_RE.search(text):
            count_falecido += 1
            continue

        # Filtro: nome obrigatório
        if opts.get("nome_obrigatorio") and not rec.get("name", "").strip():
            count_sem_nome += 1
            continue

        # Obtém o telefone principal normalizado (só dígitos)
        phone = only_digits(rec.get("phone", "") or rec.get("phone2", ""))

        # Filtro: somente com telefone
        if opts.get("somente_com_tel") and not phone:
            count_sem_tel += 1
            continue

        # ── Deduplicação: telefone EXATO ─────────────────────────────────────
        # Compara os dígitos crus, sem adicionar nem remover nada.
        # Ex: 11971234567 ≠ 1171234567  (são tratados como distintos)
        if opts.get("dup_exato") and phone:
            if phone in seen_exact:
                count_dup_exato += 1
                continue
            seen_exact.add(phone)

        # ── Deduplicação: telefone + 9º dígito ───────────────────────────────
        # Normaliza removendo o 9 central de celulares com 11 dígitos antes
        # de comparar. Isso faz com que 11971234567 e 1171234567 sejam
        # considerados o mesmo número.
        if opts.get("dup_nove") and phone:
            canon = canonical_phone(phone)   # remove 9 se tiver 11 dígitos
            if canon in seen_canonical:
                count_dup_nove += 1
                continue
            seen_canonical.add(canon)

        result.append(rec)

    # ── Relatório ────────────────────────────────────────────────────────────
    if opts.get("sem_engano"):
        print(f"  Removidos por ENGANO              : {count_engano}")
    if opts.get("sem_falecido"):
        print(f"  Removidos por FALECIDO            : {count_falecido}")
    if opts.get("nome_obrigatorio"):
        print(f"  Removidos SEM NOME                : {count_sem_nome}")
    if opts.get("somente_com_tel"):
        print(f"  Removidos SEM TELEFONE            : {count_sem_tel}")
    if opts.get("dup_exato"):
        print(f"  Removidos DUPLICADOS (exato)      : {count_dup_exato}")
    if opts.get("dup_nove"):
        print(f"  Removidos DUPLICADOS (com 9)      : {count_dup_nove}")

    return result


# ══════════════════════════════════════════════════════════════════════════════
#  ADICIONAR 9 DÍGITO NOS TELEFONES DO REGISTRO FINAL
# ══════════════════════════════════════════════════════════════════════════════

def apply_nine_digit(records: list[dict]) -> tuple[list[dict], int]:
    count = 0
    for rec in records:
        for field in ("phone", "phone2", "phone3", "phone4"):
            val = only_digits(rec.get(field, ""))
            if val:
                new_val = add_nine_digit(val)
                if new_val != val:
                    rec[field] = new_val
                    count += 1
                else:
                    rec[field] = val
    return records, count


# ══════════════════════════════════════════════════════════════════════════════
#  GERAR XLSX
# ══════════════════════════════════════════════════════════════════════════════

EXPORT_COLUMNS = [
    ("phone",        "Telefone 1"),
    ("phone2",       "Telefone 2"),
    ("phone3",       "Telefone 3"),
    ("phone4",       "Telefone 4"),
    ("ddd",          "DDD"),
    ("name",         "Nome"),
    ("cpf",          "CPF"),
    ("cnpj",         "CNPJ"),
    ("email",        "Email"),
    ("address",      "Logradouro"),
    ("number",       "Número"),
    ("complement",   "Complemento"),
    ("neighborhood", "Bairro"),
    ("city",         "Cidade"),
    ("state",        "UF"),
    ("cep",          "CEP"),
    ("purchaseDate", "Dt. Últ. Compra"),
    ("produto",      "Produto"),
    ("quantidade",   "Quantidade"),
    ("ticketAverage","Ticket Médio"),
]


def generate_xlsx(records: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Cabeçalho
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col_idx, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 18

    # Dados
    for row_idx, rec in enumerate(records, start=2):
        for col_idx, (field, _) in enumerate(EXPORT_COLUMNS, start=1):
            val = rec.get(field, "")
            if val:
                ws.cell(row=row_idx, column=col_idx, value=str(val))

    # Largura das colunas
    col_widths = {
        "Telefone 1": 16, "Telefone 2": 16, "Telefone 3": 16, "Telefone 4": 16,
        "DDD": 6, "Nome": 30, "CPF": 15, "CNPJ": 18,
        "Email": 28, "Logradouro": 28, "Número": 8, "Complemento": 14,
        "Bairro": 18, "Cidade": 18, "UF": 5, "CEP": 10,
        "Dt. Últ. Compra": 16, "Produto": 20, "Quantidade": 12, "Ticket Médio": 14,
    }
    for col_idx, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(label, 14)

    ws.freeze_panes = "A2"
    wb.save(output_path)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  DataForge — Exportação Local de Leads")
    print("=" * 60)
    print(f"\nPasta de entrada : {INPUT_DIR}")
    print(f"Pasta de saída   : {OUTPUT_DIR}\n")

    # Listar arquivos suportados na pasta input/
    supported_ext = {".csv", ".txt", ".xlsx", ".xls"}
    input_files   = sorted(
        p for p in INPUT_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in supported_ext
    )

    if not input_files:
        print(f"AVISO: Nenhum arquivo encontrado em  {INPUT_DIR}")
        print("Coloque arquivos CSV, XLSX, XLS ou TXT nessa pasta e rode novamente.")
        sys.exit(0)

    print(f"Arquivos encontrados: {len(input_files)}")
    for f in input_files:
        print(f"  - {f.name}")

    # Processar todos os arquivos
    all_records: list[dict] = []
    total_rows_read = 0

    for path in input_files:
        print(f"\nLendo: {path.name}")
        raw_rows = read_file(path)
        if not raw_rows:
            print("  (vazio ou erro — pulando)")
            continue

        print(f"  Linhas lidas      : {len(raw_rows)}")
        records = process_rows(raw_rows)
        total_rows_read += len(records)
        all_records.extend(records)
        print(f"  Registros válidos : {len(records)}")

    print(f"\nTotal lido de todos os arquivos: {total_rows_read} registros")

    if not all_records:
        print("Nenhum registro processado. Verifique os arquivos de entrada.")
        sys.exit(0)

    # Merge inteligente (CPF > CNPJ > Telefone+Nome)
    print("\nMesclando registros duplicados entre arquivos...")
    merged = merge_records(all_records)
    print(f"  Após merge: {len(merged)} registros únicos ({total_rows_read - len(merged)} mesclados)")

    # ── Perguntar filtros interativamente ─────────────────────────────────────
    opts = perguntar_filtros()

    # Filtros de exportação
    print("Aplicando filtros:")
    filtered = apply_filters(merged, opts)
    print(f"  Registros finais: {len(filtered)}")

    if not filtered:
        print("\nNenhum registro após os filtros. Planilha não gerada.")
        sys.exit(0)

    # Adicionar 9º dígito (somente se o filtro foi habilitado)
    nine_count = 0
    if opts.get("adicionar_nove"):
        filtered, nine_count = apply_nine_digit(filtered)
        print(f"  Dígito 9 adicionado em: {nine_count} telefone(s)")

    # Gerar XLSX
    timestamp   = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = OUTPUT_DIR / f"leads_{timestamp}.xlsx"

    print(f"\nGerando planilha...")
    generate_xlsx(filtered, output_path)

    print(f"\n{'='*60}")
    print(f"  Planilha gerada com sucesso!")
    print(f"  Arquivo : {output_path.name}")
    print(f"  Local   : {output_path}")
    print(f"  Total   : {len(filtered)} registros")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
